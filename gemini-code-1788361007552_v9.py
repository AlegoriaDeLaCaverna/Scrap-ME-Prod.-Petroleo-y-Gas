import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import urllib3
import io
import warnings

warnings.filterwarnings('ignore')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

st.set_page_config(page_title="Monitor de Hidrocarburos", layout="wide", page_icon="🛢️")

st.title("🛢️Monitor de Producción de Hidrocarburos🛢️")
st.markdown("Datos crudos extraídos del Ministerio de Energía, estandarizados a métricas financieras (Kbbl y KBoe).")

diccionario_fuentes = {
    'Petróleo: Producción Total': ("http://datos.energia.gob.ar/dataset/590d1284-fd6d-4686-afd8-b3da5d90a6e9/resource/4cc61040-aa44-440d-a912-91bd6c26b8a7/download/produccin-petrleo-sesco-tight-y-shale-captulo-iv-por-empresa.csv", "Petróleo Total", "m3"),
    'Petróleo: Promedio Diario': ("http://datos.energia.gob.ar/dataset/590d1284-fd6d-4686-afd8-b3da5d90a6e9/resource/2c1f455e-0103-4d51-8f94-a49c939ac0a1/download/produccin-de-petrleo-promedio-diaria-por-empresa.csv", "Petróleo Promedio", "m3/día"),
    'Gas: Producción Total': ("http://datos.energia.gob.ar/dataset/590d1284-fd6d-4686-afd8-b3da5d90a6e9/resource/63129e00-6a96-4d6e-9ce1-9e6c60287e16/download/produccin-gas-sesco-tight-y-shale-captulo-iv-por-empresa.csv", "Gas Total", "Miles de m3"),
    'Gas: Promedio Diario': ("http://datos.energia.gob.ar/dataset/590d1284-fd6d-4686-afd8-b3da5d90a6e9/resource/419094dd-2905-4ac3-9398-e81513013e5e/download/produccin-de-gas-promedio-diaria-por-empresa.csv", "Gas Promedio", "Miles de m3/día")
}

@st.cache_data(show_spinner=False)
def cargar_datos(url):
    try:
        df = pd.read_csv(url, encoding='utf-8-sig', sep=None, engine='python')
    except UnicodeDecodeError:
        df = pd.read_csv(url, encoding='latin-1', sep=None, engine='python')

    df.columns = df.columns.str.lower().str.strip().str.replace('ï»¿', '')

    for col in df.columns:
        if 'empresa' in col or 'operador' in col:
            df.rename(columns={col: 'empresa'}, inplace=True)
        elif 'anio' in col or 'año' in col or 'year' in col:
            df.rename(columns={col: 'anio'}, inplace=True)
        elif col in ['mes', 'month']:
            df.rename(columns={col: 'mes'}, inplace=True)

    cols_base = ['empresa', 'anio', 'mes']
    faltantes = [c for c in cols_base if c not in df.columns]
    if faltantes:
        st.error(f"El Estado modificó las columnas. Faltan: {faltantes}")
        st.stop()

    columnas_prohibidas = ['indice_tiempo', 'id']
    cols_metricas = [c for c in df.columns if c not in cols_base and c not in columnas_prohibidas and df[c].dtype in [np.float64, np.int64]]

    if not cols_metricas:
        st.error("No hay columnas numéricas de volumen en este archivo.")
        st.stop()

    df = df.dropna(subset=cols_base)
    df['anio'] = pd.to_numeric(df['anio'], errors='coerce').fillna(0).astype(int)
    df['mes'] = pd.to_numeric(df['mes'], errors='coerce').fillna(0).astype(int)
    df = df[df['anio'] > 0]
    
    return df, cols_metricas

col1, col2 = st.columns([2, 1])

with col1:
    seleccion_fuente = st.selectbox("Seleccione el origen de datos:", list(diccionario_fuentes.keys()))

url_csv, nombre_fluido, unidad_medida = diccionario_fuentes[seleccion_fuente]

es_petroleo = "Petróleo" in nombre_fluido
es_promedio = "Promedio" in nombre_fluido
es_desglosable = "Total" in seleccion_fuente

with st.spinner('⏳Extrayendo pergaminos del Ministerio...⏳'):
    df, cols_metricas = cargar_datos(url_csv)

anios_disponibles = sorted(df['anio'].unique().tolist())
lista_empresas = ['(Todas)'] + sorted(df['empresa'].unique().tolist())

with col2:
    anio_desde, anio_hasta = st.slider("📏Rango de Años:", min_value=min(anios_disponibles), max_value=max(anios_disponibles), value=(min(anios_disponibles), max(anios_disponibles)))
    empresa_elegida = st.selectbox("🏭Seleccione una Empresa:", lista_empresas)
    
    if es_desglosable:
        tipo_grafico = st.radio("📈 Variables a graficar:", ["Total Consolidado", "Convencional vs No Convencional"])
    else:
        tipo_grafico = "Total Consolidado"

if 'procesar_clicked' not in st.session_state:
    st.session_state.procesar_clicked = False

if st.button("🎯Procesar y Previsualizar", type="primary"):
    st.session_state.procesar_clicked = True

if st.session_state.procesar_clicked:
    
    df_filtrado = df[(df['anio'] >= anio_desde) & (df['anio'] <= anio_hasta)].copy()
    if empresa_elegida != '(Todas)':
        df_filtrado = df_filtrado[df_filtrado['empresa'] == empresa_elegida]
        
    if df_filtrado.empty:
        st.warning("El pozo está seco. No hay registros para esta combinación.")
        st.stop()
        
    df_filtrado['Periodo'] = df_filtrado['anio'].astype(str) + "-" + df_filtrado['mes'].astype(str).str.zfill(2)
    
    se_puede_desglosar = 'concepto' in df_filtrado.columns
    
    if se_puede_desglosar:
        col_valor = cols_metricas[0] 
        df_agrupado = df_filtrado.pivot_table(index='Periodo', columns='concepto', values=col_valor, aggfunc='sum').reset_index()
        df_agrupado = df_agrupado.fillna(0)
        cols_grafico = [c for c in df_agrupado.columns if c != 'Periodo']
    else:
        df_agrupado = df_filtrado.groupby('Periodo')[cols_metricas].sum().reset_index()
        cols_grafico = cols_metricas
        
    df_agrupado = df_agrupado.sort_values('Periodo')
    
    # --- LA INGENIERÍA FINANCIERA (CORREGIDA) ---
    factor_conversion = 6.2898 / 1000

    if es_petroleo:
        unidad_final = "Miles de Barriles/día" if es_promedio else "Miles de Barriles"
        acronimo_final = "Kbbl/d" if es_promedio else "Kbbl"
    else:
        unidad_final = "Miles de BOE/día" if es_promedio else "Miles de BOE"
        acronimo_final = "KBoe/d" if es_promedio else "KBoe"

    nombre_col_nueva = f"{'Promedio' if es_promedio else 'Total'} Consolidado ({acronimo_final})"
    
    # 1. Calculamos la columna matriz antes que cualquier otra cosa
    suma_cruda = df_agrupado[cols_grafico].sum(axis=1)
    df_agrupado[nombre_col_nueva] = (suma_cruda * factor_conversion).round(2)
    
    # 2. Ejecutamos la bifurcación geológica
    if se_puede_desglosar:
        cols_no_conv_raw = [c for c in cols_grafico if any(k in str(c).lower() for k in ['shale', 'tight', 'no convencional'])]
        cols_conv_raw = [c for c in cols_grafico if c not in cols_no_conv_raw]
        
        raw_no_conv = df_agrupado[cols_no_conv_raw].sum(axis=1) if cols_no_conv_raw else 0
        raw_conv = df_agrupado[cols_conv_raw].sum(axis=1) if cols_conv_raw else 0
        
        col_conv_convrt = f"Convencional ({acronimo_final})"
        col_noconv_convrt = f"Total No Convencional ({acronimo_final})"
        
        df_agrupado[col_conv_convrt] = (raw_conv * factor_conversion).round(2)
        df_agrupado[col_noconv_convrt] = (raw_no_conv * factor_conversion).round(2)
        
        cols_to_graph_web = [col_conv_convrt, col_noconv_convrt] if tipo_grafico == "Convencional vs No Convencional" else [nombre_col_nueva]
        
        columnas_finales = ['Periodo'] + cols_grafico + [col_conv_convrt, col_noconv_convrt, nombre_col_nueva]
        df_agrupado = df_agrupado[columnas_finales]
    else:
        cols_to_graph_web = [nombre_col_nueva]
        columnas_finales = ['Periodo'] + cols_grafico + [nombre_col_nueva]
        df_agrupado = df_agrupado[columnas_finales]

    # --- VITRINA DE PREVISUALIZACIÓN WEB ---
    st.markdown("---")
    st.subheader(f"📊Vista Previa: {nombre_fluido} - {empresa_elegida}")
    
    df_grafico_web = df_agrupado.set_index('Periodo')[cols_to_graph_web]
    st.line_chart(df_grafico_web, y_label=unidad_final)
    st.dataframe(df_agrupado, use_container_width=True)

    # --- CREACIÓN DEL EXCEL EN MEMORIA RAM ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {nombre_fluido}"
    
    ws.append(list(df_agrupado.columns))
    
    for r in dataframe_to_rows(df_agrupado, index=False, header=False):
        ws.append(r)
        
    fila_notas = ws.max_row + 2
    ws.cell(row=fila_notas, column=1, value="NOTAS Y FUENTES METODOLÓGICAS:").font = Font(bold=True, italic=True)
    ws.cell(row=fila_notas + 1, column=1, value="• Base de datos original: Producción de Petróleo y Gas (SESCO), Secretaría de Energía, Ministerio de Economía.")
    ws.cell(row=fila_notas + 2, column=1, value="• Factores de conversión: Tabla de Conversiones, Pampa Energía.")
    
    nota_divergencia = "• Aclaración de titularidad: La divergencia entre los registros estatales y los balances corporativos radica en el criterio de imputación. Las fuentes oficiales asignan el 100% de la extracción al operador técnico del área, mientras que las empresas reportan su producción según su porcentaje de participación societaria (Working Interest)."
    ws.cell(row=fila_notas + 3, column=1, value=nota_divergencia)
    
    if es_petroleo:
        texto_calculo = "• Cálculo Petróleo: Se multiplican los metros cúbicos (m3) por 6,2898 y se dividen por 1.000 para obtener Miles de Barriles (Kbbl)."
    else:
        texto_calculo = "• Cálculo Gas: Se trata el dato reportado en Mm3 como equivalente directo a m3 de petróleo. Se multiplica por 6,2898 y se divide por 1.000 para obtener Miles de Barriles Equivalentes de Petróleo (KBoe)."
        
    ws.cell(row=fila_notas + 4, column=1, value=texto_calculo)
    
    # --- CONSTRUCCIÓN DEL GRÁFICO DE EXCEL ---
    chart = LineChart()
    chart.title = f"{nombre_fluido} - {empresa_elegida} ({acronimo_final})"
    chart.style = 13
    chart.width = 24  
    chart.height = 12 
    chart.x_axis.title = "Período"
    chart.y_axis.title = f"Volumen ({unidad_final})"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.numFmt = '#,##0'
    
    if se_puede_desglosar and tipo_grafico == "Convencional vs No Convencional":
        min_c = len(columnas_finales) - 2 
        max_c = len(columnas_finales) - 1
        chart.legend.position = 'b'
    else:
        min_c = len(columnas_finales)
        max_c = len(columnas_finales)
        chart.legend = None 
        
    datos_ref = Reference(ws, min_col=min_c, min_row=1, max_col=max_c, max_row=ws.max_row - 5)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row - 5)
    
    chart.add_data(datos_ref, titles_from_data=True, from_rows=False)
    chart.set_categories(cat_ref)
    
    ws.add_chart(chart, "H2")
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    fluido_archivo = "Petroleo" if "Petróleo" in nombre_fluido else "Gas"
    tipo_archivo = "Total" if "Total" in nombre_fluido else "Promedio"
    nombre_empresa_limpio = empresa_elegida.replace(' ', '_').replace('/', '-')
    nombre_archivo = f"Reporte_{fluido_archivo}_{tipo_archivo}_{nombre_empresa_limpio}.xlsx"
    
    st.markdown("---")
    st.download_button(
        label="👨‍💻Descargar Archivo Excel",
        data=buffer,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
