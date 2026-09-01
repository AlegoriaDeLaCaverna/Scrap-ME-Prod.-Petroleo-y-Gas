import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import urllib3
import io
import warnings

# Silenciamos las alarmas del Estado
warnings.filterwarnings('ignore')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

st.set_page_config(page_title="Monitor de Actividad Empresarial", layout="wide", page_icon="📈")

st.title("Monitor de Producción y Ventas (Actividad por Empresa)")
st.markdown("Datos crudos extraídos de infra.datos.gob.ar. Formato de series de tiempo consolidadas.")

# Catálogo de los nuevos pergaminos
diccionario_urls = {
    'Frecuencia 1 (365.1)': "https://infra.datos.gob.ar/catalog/sspm/dataset/365/distribution/365.1/download/actividad-empresas.csv",
    'Frecuencia 2 (365.2)': "https://infra.datos.gob.ar/catalog/sspm/dataset/365/distribution/365.2/download/actividad-empresas.csv",
    'Frecuencia 3 (365.3)': "https://infra.datos.gob.ar/catalog/sspm/dataset/365/distribution/365.3/download/actividad-empresas.csv"
}

# --- MOTOR DE INGESTA ---
@st.cache_data(show_spinner=False)
def cargar_datos(url):
    try:
        df = pd.read_csv(url, encoding='utf-8-sig')
    except Exception:
        df = pd.read_csv(url, encoding='latin-1')

    # Identificar la columna de tiempo (suele llamarse indice_tiempo)
    col_tiempo = 'indice_tiempo' if 'indice_tiempo' in df.columns else df.columns[0]
    
    # Formatear la fecha para que el gráfico de Excel no colapse
    df[col_tiempo] = pd.to_datetime(df[col_tiempo], errors='coerce')
    df = df.dropna(subset=[col_tiempo])
    df[col_tiempo] = df[col_tiempo].dt.strftime('%Y-%m') # Formato estricto Año-Mes
    
    # El resto son todas métricas
    cols_metricas = [c for c in df.columns if c != col_tiempo]
    
    return df, col_tiempo, cols_metricas

# --- INTERFAZ WEB ---
col1, col2 = st.columns([1, 2])

with col1:
    seleccion_fuente = st.selectbox("Seleccione el set de datos a auditar:", list(diccionario_urls.keys()))

url_csv = diccionario_urls[seleccion_fuente]

with st.spinner('Extrayendo los expedientes...'):
    df, col_tiempo, cols_metricas = cargar_datos(url_csv)

with col2:
    # Como los títulos son largos, damos un multiselect para que el usuario elija qué curvas graficar
    metricas_elegidas = st.multiselect(
        "Seleccione las series a incluir en el reporte (Puede elegir varias):", 
        cols_metricas,
        default=[cols_metricas[0]] if cols_metricas else None
    )

if st.button("Procesar Datos y Generar Excel", type="primary"):
    
    if not metricas_elegidas:
        st.warning("Debe seleccionar al menos una serie para generar el reporte.")
        st.stop()
        
    # Filtramos la tabla solo con el tiempo y las métricas elegidas
    columnas_finales = [col_tiempo] + metricas_elegidas
    df_filtrado = df[columnas_finales].copy()
    
    # Reemplazamos los NaN por vacío para que Excel no grafique caídas a cero inexistentes
    df_filtrado = df_filtrado.replace({np.nan: None})
    
    # --- CREACIÓN DEL EXCEL EN MEMORIA RAM ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividad Empresas"
    
    # Escribimos los encabezados (limpiando un poco los guiones bajos para estética)
    encabezados = ['Período'] + [m.replace('_', ' ').title() for m in metricas_elegidas]
    ws.append(encabezados)
    
    for r in dataframe_to_rows(df_filtrado, index=False, header=False):
        ws.append(r)
        
    chart = LineChart()
    chart.title = "Evolución de Actividad"
    chart.style = 13
    chart.width = 24  
    chart.height = 12 
    
    chart.x_axis.title = "Período"
    chart.y_axis.title = "Volumen / Cantidad"
    
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.numFmt = '#,##0'
    
    # Referenciamos los datos
    datos_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(metricas_elegidas), max_row=ws.max_row)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    
    chart.add_data(datos_ref, titles_from_data=True, from_rows=False)
    chart.set_categories(cat_ref)
    
    if len(metricas_elegidas) == 1:
        chart.legend = None 
    else:
        chart.legend.position = 'b' 
    
    ws.add_chart(chart, "F2")
    
    # Guardar en memoria virtual
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    nombre_archivo = "Reporte_Actividad_Empresas.xlsx"
    
    st.success("¡El reporte ha sido forjado con éxito!")
    
    st.download_button(
        label="📥 Descargar Archivo Excel",
        data=buffer,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )